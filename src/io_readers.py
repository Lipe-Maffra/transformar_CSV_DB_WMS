from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Callable, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

EXPECTED_WMS_COLUMNS = [
    "Tarefa",
    "Tipo",
    "Onda",
    "Data inicial",
    "Data final",
    "Item",
    "Origem",
    "Destino",
    "Qtd.",
    "Emb.",
    "Exec.",
    "Prio",
    "Lib.",
    "Tar. bloq.",
    "Prio.B",
    "Bloqs.",
    "Pedido",
    "Aut. receb.",
    "Data tarefa",
    "Lote",
    "Almoxarifado",
    "Carga",
    "Unitizador",
]

TECH_COLUMNS = [
    "_source_file",
    "_source_sheet",
    "_source_mtime",
    "_loaded_at",
]

_ISO_TS_RE = re.compile(r"\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}(:\d{2})?")
_WF_HEADER_RE = re.compile(r"wf0008|manuten[cç][aã]o de tarefas", re.IGNORECASE)
_JUNK_COL_RE = re.compile(r"^resultado$", re.IGNORECASE)


def _normalize_col_name(name: object) -> str:
    text = "" if name is None else str(name)
    text = text.replace("\ufeff", "").strip()
    text = " ".join(text.split())
    return text


def _normalize_key(text: str) -> str:
    text = text.lower()
    text = text.replace(".", "").replace(" ", "")
    return text


def _is_bad_header_cell(value: object) -> bool:
    text = _normalize_col_name(value)
    if not text:
        return False
    if ":\\" in text:
        return True
    if _ISO_TS_RE.search(text):
        return True
    return False


def _detect_csv_header(path: Path) -> Tuple[int, str]:
    encodings = ["utf-8-sig", "utf-8", "latin-1"]
    header_row = 0

    for enc in encodings:
        try:
            with path.open("r", encoding=enc, errors="ignore") as fh:
                first_line = fh.readline()
            if first_line and _WF_HEADER_RE.search(first_line):
                header_row = 1
            break
        except Exception:
            continue

    delim = ";"
    try:
        df_probe = pd.read_csv(
            path,
            sep=delim,
            encoding=encodings[0],
            dtype=str,
            engine="python",
            keep_default_na=False,
            na_values=[],
            on_bad_lines="skip",
            header=header_row,
            nrows=1,
        )
        if df_probe.shape[1] <= 1:
            delim = ","
    except Exception:
        delim = ";"

    return header_row, delim


def _detect_excel_header(path: Path) -> Tuple[int, Optional[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        a1 = ws["A1"].value
        if a1 and _WF_HEADER_RE.search(str(a1)):
            return 1, ws.title
        return 0, ws.title
    finally:
        wb.close()


def detect_header_row(path: Path) -> Tuple[int, Optional[str], Optional[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        header_row, delim = _detect_csv_header(path)
        return header_row, delim, None

    if suffix in (".xlsx", ".xlsm"):
        header_row, sheet_name = _detect_excel_header(path)
        return header_row, None, sheet_name

    if suffix == ".xls":
        try:
            header_row, sheet_name = _detect_excel_header(path)
            return header_row, None, sheet_name
        except Exception:
            return 0, None, None

    return 0, None, None


def _is_junk_column(name: str) -> bool:
    if not name:
        return True
    lower = name.lower()
    if lower.startswith("unnamed"):
        return True
    if _normalize_key(name).startswith("nan"):
        return True
    if _WF_HEADER_RE.search(name):
        return True
    if _JUNK_COL_RE.match(name):
        return True
    if ":\\" in name:
        return True
    if _ISO_TS_RE.search(name):
        return True
    return False


def _map_special_columns(name: str) -> str:
    lower = name.lower()
    key = _normalize_key(name)
    if key in {"sourcefile", "source_file", "arquivoorigem", "arquivo_origem"}:
        return "_source_file"
    if lower.endswith("ile") and ("source" in lower or "arquivo" in lower):
        return "_source_file"
    return name


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    expected_map = {_normalize_key(name): name for name in EXPECTED_WMS_COLUMNS}
    tech_map = {_normalize_key(name): name for name in TECH_COLUMNS}
    new_cols: list[str] = []
    keep_mask: list[bool] = []

    for c in df.columns:
        norm = _normalize_col_name(c)
        if not norm:
            keep_mask.append(False)
            continue
        norm = _map_special_columns(norm)
        norm_key = _normalize_key(norm)

        if _is_junk_column(norm):
            keep_mask.append(False)
            continue

        if norm_key in expected_map:
            new_cols.append(expected_map[norm_key])
            keep_mask.append(True)
            continue
        if norm_key in tech_map:
            new_cols.append(tech_map[norm_key])
            keep_mask.append(True)
            continue

        keep_mask.append(False)

    df = df.loc[:, keep_mask].copy()
    df.columns = new_cols

    seen: set[str] = set()
    dedup_cols: list[str] = []
    dedup_mask: list[bool] = []
    for c in df.columns:
        if c in seen:
            dedup_mask.append(False)
            continue
        seen.add(c)
        dedup_cols.append(c)
        dedup_mask.append(True)

    df = df.loc[:, dedup_mask].copy()
    df.columns = dedup_cols
    return df


def _drop_fully_empty_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    tmp = df.replace(r"^\s*$", pd.NA, regex=True)
    return df.loc[~tmp.isna().all(axis=1)].reset_index(drop=True)


def _ensure_schema(df: pd.DataFrame) -> pd.DataFrame:
    for c in EXPECTED_WMS_COLUMNS:
        if c not in df.columns:
            df[c] = pd.NA

    for c in TECH_COLUMNS:
        if c not in df.columns:
            df[c] = pd.NA

    ordered = EXPECTED_WMS_COLUMNS + TECH_COLUMNS
    return df.loc[:, ordered]


def _ensure_required_columns(df: pd.DataFrame, path: Path) -> None:
    if "Tarefa" not in df.columns:
        raise ValueError(
            f"Required column 'Tarefa' not found after read: {path.name}. "
            f"Columns: {list(df.columns)}"
        )


def _read_csv_with_fallback(path: Path, header_row: int, delim: str) -> pd.DataFrame:
    encodings = ["utf-8-sig", "utf-8", "latin-1"]
    last_err: Optional[Exception] = None
    for enc in encodings:
        try:
            return pd.read_csv(
                path,
                sep=delim,
                encoding=enc,
                dtype=str,
                engine="python",
                keep_default_na=False,
                na_values=[],
                on_bad_lines="skip",
                header=header_row,
            )
        except Exception as e:
            last_err = e
    raise ValueError(f"Failed to read CSV: {path.name}. Last error: {last_err}")


def _read_excel_with_header(path: Path, header_row: int, sheet_name: Optional[str]) -> pd.DataFrame:
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        return pd.read_excel(
            path,
            dtype=str,
            engine="openpyxl",
            header=header_row,
            sheet_name=sheet_name,
        )
    return pd.read_excel(path, dtype=str, header=header_row, sheet_name=sheet_name)


def _clean_date_columns(df: pd.DataFrame) -> None:
    for col in ("Data inicial", "Data final", "Data tarefa"):
        if col in df.columns:
            df[col] = df[col].apply(lambda v: v.strip() if isinstance(v, str) else v)


def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = _normalize_columns(df)
    df = _drop_fully_empty_rows(df)
    df = _ensure_schema(df)
    _clean_date_columns(df)
    return df


def read_data_file(path: Path, log: Optional[Callable[[str], None]] = None) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    suffix = path.suffix.lower()
    header_row, delim, sheet_name = detect_header_row(path)

    if suffix == ".csv":
        df = _read_csv_with_fallback(path, header_row, delim or ";")
        df = _clean_dataframe(df)
        _ensure_required_columns(df, path)
        if log:
            log(
                f"[csv] file={path} header_row={header_row} delim={repr(delim)} "
                f"cols={list(df.columns)} rows={len(df)}"
            )
        return df

    if suffix in (".xlsx", ".xlsm", ".xls"):
        df = _read_excel_with_header(path, header_row, sheet_name)
        df = _clean_dataframe(df)
        _ensure_required_columns(df, path)
        df.attrs["source_sheet"] = sheet_name
        if log:
            log(
                f"[excel] file={path} header_row={header_row} sheet={sheet_name} "
                f"cols={list(df.columns)} rows={len(df)}"
            )
        return df

    raise ValueError(f"Unsupported extension: {path.name}")


def list_data_files(folder: Path, recursive: bool = True) -> list[Path]:
    """
    Lista arquivos suportados.
    - recursive=True: inclui subpastas (rglob)
    """
    if not folder.exists():
        raise FileNotFoundError(f"Folder not found: {folder}")

    files: list[Path] = []
    for ext in ("*.csv", "*.xlsx", "*.xlsm", "*.xls"):
        if recursive:
            files.extend(folder.rglob(ext))
        else:
            files.extend(folder.glob(ext))

    # ignore Excel temp files (~$)
    files = [p for p in files if p.is_file() and not p.name.startswith("~$")]

    return sorted(files, key=lambda p: str(p).lower())
